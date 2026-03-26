import openpyxl

wb = openpyxl.load_workbook('2025년_중3_1학기 기말_봉황중학교_충청남도 공주시_동아(윤정미).xlsx')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"Sheet: {sn}, Rows: {ws.max_row}, Cols: {ws.max_column}")
    # Print merged cells
    if ws.merged_cells.ranges:
        merges = list(ws.merged_cells.ranges)
        print(f"  Merged cells ({len(merges)}): {merges[:20]}")
    # Print header row
    headers = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        headers.append(str(cell.value))
    print(f"  Headers: {headers}")
    # Print all rows
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if val is not None:
                val_str = repr(val)
                if len(val_str) > 200:
                    val_str = val_str[:200] + "..."
                row_data[headers[c-1]] = val_str
        if row_data:
            print(f"  Row {r}: {row_data}")

