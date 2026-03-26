import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('2025년_중3_1학기 기말_봉황중학교_충청남도 공주시_동아(윤정미).xlsx')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"Sheet: {sn}, Rows: {ws.max_row}, Cols: {ws.max_column}")
    if ws.merged_cells.ranges:
        merges = list(ws.merged_cells.ranges)
        print(f"  Merged cells ({len(merges)}): {merges}")
    headers = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        headers.append(str(cell.value))
    print(f"  Headers: {headers}")
    for r in range(2, ws.max_row + 1):
        print(f"\n--- Row {r} ---")
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if val is not None:
                print(f"  {headers[c-1]}: {repr(val)}")

